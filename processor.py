"""
CSV parsing and XLSX building for the Entri bulk sharing-link generator.
"""

import copy
import csv
import io
import logging
import re
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from entri_client import EntriClient, EntriError

logger = logging.getLogger(__name__)


# Permissive domain pattern: labels separated by dots, must contain at least one dot.
# Real validation happens at Entri; this just guards against obvious junk rows.
_DOMAIN_RE = re.compile(
    r"^(?=.{1,253}$)(?!-)[A-Za-z0-9-]{1,63}(?<!-)(\.(?!-)[A-Za-z0-9-]{1,63}(?<!-))+$"
)


def parse_domains_from_csv(raw: bytes) -> List[str]:
    """
    Extract a list of domains from CSV bytes.

    Accepts:
      - A single column of domains, with or without a header.
      - A column literally named 'domain' (case-insensitive) anywhere in the file.

    Raises ValueError on malformed input.
    """
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = raw.decode("latin-1")
        except UnicodeDecodeError as exc:
            raise ValueError(f"could not decode file as text: {exc}") from exc

    # Sniff the dialect; fall back to default comma if sniffing fails.
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect=dialect)
    rows = [r for r in reader if any(cell.strip() for cell in r)]

    if not rows:
        return []

    # If first row looks like a header containing 'domain', use that column.
    first_row = [c.strip() for c in rows[0]]
    domain_col_idx = None
    for i, cell in enumerate(first_row):
        if cell.lower() == "domain":
            domain_col_idx = i
            break

    if domain_col_idx is not None:
        data_rows = rows[1:]
    else:
        # No header row identified; assume column 0 is the domain column.
        data_rows = rows
        domain_col_idx = 0

    domains: List[str] = []
    seen = set()
    for r in data_rows:
        if domain_col_idx >= len(r):
            continue
        candidate = r[domain_col_idx].strip().lower()
        if not candidate:
            continue
        # Strip common URL prefixes if user pasted full URLs.
        candidate = re.sub(r"^https?://", "", candidate)
        candidate = candidate.split("/", 1)[0]
        if candidate in seen:
            continue
        seen.add(candidate)
        domains.append(candidate)

    return domains


def process_domains_to_xlsx(
    client: EntriClient,
    domains: List[str],
    sharing_flow: str,
    base_config: Dict[str, Any],
) -> bytes:
    """
    Generate a sharing link for each domain and pack the results into XLSX.

    Returns the workbook serialized to bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sharing Links"

    # --- Header row ---
    headers = ["Domain", "Sharing Link", "Status"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="left", vertical="center")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Body ---
    for i, domain in enumerate(domains, start=1):
        # Validate locally so we don't waste API calls on garbage.
        if not _DOMAIN_RE.match(domain):
            ws.append([domain, "", "ERROR: invalid domain format"])
            logger.warning("Skipped invalid domain: %s", domain)
            continue

        # Build a config copy for this domain.
        per_domain_config = copy.deepcopy(base_config)
        per_domain_config["prefilledDomain"] = domain

        try:
            data = client.create_sharing_link(per_domain_config, flow=sharing_flow)
            link = data.get("link", "")
            ws.append([domain, link, "OK"])
            logger.info("[%d/%d] %s -> %s", i, len(domains), domain, link)
        except EntriError as exc:
            err_msg = str(exc)
            # Keep the cell readable.
            if len(err_msg) > 250:
                err_msg = err_msg[:247] + "..."
            ws.append([domain, "", f"ERROR: {err_msg}"])
            logger.warning("[%d/%d] %s failed: %s", i, len(domains), domain, exc)

    # --- Column widths ---
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 40

    # Freeze header row.
    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
